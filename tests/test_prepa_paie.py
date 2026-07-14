"""
Tests de la page Preparation de paie :
- controle d'acces
- rendu de la grille : colonnes figees (Traite, Nom Prenom) presentes,
  colonne "Nom Prenom" dupliquee en fin de tableau supprimee
"""


def _inserer_contrat(db, user_id):
    """Cree un contrat actif pour que le salarie apparaisse dans la grille."""
    db.execute(
        "INSERT INTO contrats (user_id, type_contrat, date_debut, date_fin) "
        "VALUES (?, 'CDI', '2026-01-01', NULL)", (user_id,)
    )
    db.commit()


def _inserer_contrat_cdd(db, user_id, temps_hebdo=28.0):
    """Cree un CDD actif avec un temps de travail hebdomadaire."""
    db.execute(
        "INSERT INTO contrats (user_id, type_contrat, date_debut, date_fin, temps_hebdo) "
        "VALUES (?, 'CDD', '2026-01-01', '2026-12-31', ?)", (user_id, temps_hebdo)
    )
    db.commit()


class TestPagePrepaPaie:

    def test_acces_refuse_salarie(self, auth_client):
        resp = auth_client.get('/prepa_paie', follow_redirects=False)
        assert resp.status_code == 302

    def test_grille_colonnes_figees_sans_doublon_nom(self, comptable_client, sample_users, app, db):
        with app.app_context():
            _inserer_contrat(db, sample_users['salarie_id'])

        resp = comptable_client.get('/prepa_paie?mois=6&annee=2026')
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')

        # La grille est bien rendue avec le salarie sous contrat
        assert f"traite_{sample_users['salarie_id']}" in html
        # Colonnes figees a gauche
        assert 'pp-col-traite' in html
        assert 'pp-col-nom' in html
        # L'en-tete "Nom Prenom" n'apparait qu'une fois : la colonne dupliquee
        # en fin de tableau a ete supprimee
        assert html.count('>Nom Prenom</th>') == 1
        # Le nom du salarie n'apparait qu'une fois par ligne (plus de rappel a droite)
        assert html.count('<strong>Martin</strong>') == 1

    def test_cdd_affiche_temps_hebdo_entre_parentheses(self, comptable_client, sample_users, app, db):
        """Pour un CDD, le temps de travail hebdomadaire suit les dates entre parentheses."""
        with app.app_context():
            _inserer_contrat_cdd(db, sample_users['salarie_id'], temps_hebdo=28.0)

        resp = comptable_client.get('/prepa_paie?mois=6&annee=2026')
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')

        # Les dates du contrat sont suivies du temps hebdo entre parentheses
        assert '(28.0h)' in html

    def test_cdi_n_affiche_pas_temps_hebdo(self, comptable_client, sample_users, app, db):
        """Le temps hebdo entre parentheses est reserve aux CDD (pas les CDI)."""
        with app.app_context():
            db.execute(
                "INSERT INTO contrats (user_id, type_contrat, date_debut, date_fin, temps_hebdo) "
                "VALUES (?, 'CDI', '2026-01-01', NULL, 35.0)",
                (sample_users['salarie_id'],)
            )
            db.commit()

        resp = comptable_client.get('/prepa_paie?mois=6&annee=2026')
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')

        # Pas de temps hebdo entre parentheses pour un CDI
        assert '(35.0h)' not in html

    def test_colonne_pdf_separee_du_contrat(self, comptable_client, sample_users, app, db):
        """Le lien PDF du contrat dispose de sa propre colonne, apres 'Contrat'."""
        with app.app_context():
            _inserer_contrat_cdd(db, sample_users['salarie_id'])

        resp = comptable_client.get('/prepa_paie?mois=6&annee=2026')
        assert resp.status_code == 200
        html = resp.data.decode('utf-8')

        # En-tete dedie 'PDF' present, place juste apres l'en-tete 'Contrat'
        assert '>PDF</th>' in html
        assert html.index('>Contrat</th>') < html.index('>PDF</th>') < html.index('>Forfait</th>')

    def test_export_excel_contient_temps_hebdo(self, comptable_client, sample_users, app, db):
        """L'export Excel comporte une colonne 'Temps Hebdo (h)' renseignee."""
        from io import BytesIO
        from openpyxl import load_workbook

        with app.app_context():
            _inserer_contrat_cdd(db, sample_users['salarie_id'], temps_hebdo=28.0)

        resp = comptable_client.get('/prepa_paie/export_excel?mois=6&annee=2026')
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.data))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert 'Temps Hebdo (h)' in headers
        # La colonne est positionnee juste apres 'Date fin'
        assert headers.index('Temps Hebdo (h)') == headers.index('Date fin') + 1
        # La valeur du CDD figure sur la ligne du salarie (ligne 2)
        col = headers.index('Temps Hebdo (h)') + 1
        assert ws.cell(row=2, column=col).value == 28.0

    def test_variables_du_mois_sans_contrat_actif_apparait(self, comptable_client,
                                                           sample_users, app, db):
        """CDD terminé le mois précédent, solde d'heures supp payé sur la paie
        du mois suivant : la salariée doit apparaître en prépa paie (sinon le
        prestataire ne voit jamais ces heures), même sans contrat actif."""
        uid = sample_users['salarie_id']
        db.execute("INSERT INTO contrats (user_id, type_contrat, date_debut, date_fin) "
                   "VALUES (?, 'CDD', '2026-01-01', '2026-06-30')", (uid,))
        db.execute("INSERT INTO variables_paie (user_id, mois, annee, heures_supps, "
                   "hs_deduites_compteur) VALUES (?, 7, 2026, 6, 1)", (uid,))
        db.commit()

        html = comptable_client.get('/prepa_paie?mois=7&annee=2026').get_data(as_text=True)
        assert f"traite_{uid}" in html          # présente malgré le contrat terminé
        assert '6' in html                      # ses heures supp sont visibles
        assert 'Aucun contrat actif ce mois' in html   # mention explicite

    def test_variables_vides_sans_contrat_actif_reste_absent(self, comptable_client,
                                                             sample_users, app, db):
        """Une ligne de variables « vide » (reports persistants type mutuelle
        uniquement) ne fait pas réapparaître un salarié sans contrat actif :
        seul un montant/des heures/un commentaire du mois compte."""
        uid = sample_users['salarie_id']
        db.execute("INSERT INTO contrats (user_id, type_contrat, date_debut, date_fin) "
                   "VALUES (?, 'CDD', '2026-01-01', '2026-06-30')", (uid,))
        db.execute("INSERT INTO variables_paie (user_id, mois, annee, mutuelle, "
                   "hs_deduites_compteur) VALUES (?, 7, 2026, 1, 1)", (uid,))
        db.commit()

        html = comptable_client.get('/prepa_paie?mois=7&annee=2026').get_data(as_text=True)
        assert f"traite_{uid}" not in html

    def test_export_excel_inclut_variables_sans_contrat(self, comptable_client,
                                                        sample_users, app, db):
        """L'export Excel suit la même règle que la page."""
        from io import BytesIO
        from openpyxl import load_workbook
        uid = sample_users['salarie_id']
        db.execute("INSERT INTO contrats (user_id, type_contrat, date_debut, date_fin) "
                   "VALUES (?, 'CDD', '2026-01-01', '2026-06-30')", (uid,))
        db.execute("INSERT INTO variables_paie (user_id, mois, annee, heures_supps, "
                   "hs_deduites_compteur) VALUES (?, 7, 2026, 6, 1)", (uid,))
        db.commit()

        resp = comptable_client.get('/prepa_paie/export_excel?mois=7&annee=2026')
        assert resp.status_code == 200
        ws = load_workbook(BytesIO(resp.data)).active
        noms = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert any(n and 'Martin' in n for n in noms)
