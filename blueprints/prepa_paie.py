"""
Blueprint prepa_paie_bp.
Page de preparation de paie mensuelle : synthese de toutes les infos
necessaires pour le traitement de la paie de chaque salarie.
Accessible par comptable, directeur et prestataire.
"""
from io import BytesIO
from flask import (Blueprint, render_template, request, redirect,
                   url_for, session, flash, make_response)
from datetime import datetime
from database import get_db
from utils import login_required, NOMS_MOIS

prepa_paie_bp = Blueprint('prepa_paie_bp', __name__)

# Motifs d'absence affiches dans la prepa paie (hors recuperations)
MOTIFS_ABSENCE_PAIE = [
    'Arrêt maladie',
    'Congé payé',
    'Congé conventionnel',
    'Congé parental',
    'Jour enfant malade',
    'Accident du travail',
    'Evènement familial',
    'Sans solde',
    'Mi-temps thérapeutique',
    'Autre',
]


def _peut_acceder_prepa_paie():
    """Comptable, directeur et prestataire peuvent acceder a cette page."""
    return session.get('profil') in ['comptable', 'directeur', 'prestataire']


def _get_salaries_avec_contrat_actif(conn, mois, annee):
    """Retourne les salaries ayant (ou ayant eu) un contrat actif durant le mois."""
    # Calculer les bornes du mois
    date_debut_mois = f"{annee:04d}-{mois:02d}-01"
    if mois == 12:
        date_fin_mois = f"{annee:04d}-12-31"
    else:
        # Dernier jour du mois : premier jour du mois suivant - 1
        from calendar import monthrange
        _, dernier_jour = monthrange(annee, mois)
        date_fin_mois = f"{annee:04d}-{mois:02d}-{dernier_jour:02d}"

    # Un salarie a un contrat actif sur le mois si :
    # - date_debut <= date_fin_mois ET (date_fin IS NULL OR date_fin >= date_debut_mois)
    salaries = conn.execute('''
        SELECT DISTINCT u.id, u.nom, u.prenom, u.profil,
               COALESCE(s.nom, '') AS secteur_nom
        FROM users u
        JOIN contrats c ON c.user_id = u.id
        LEFT JOIN secteurs s ON u.secteur_id = s.id
        WHERE u.profil != 'prestataire'
        AND c.date_debut <= %s
        AND (c.date_fin IS NULL OR c.date_fin >= %s)
        ORDER BY u.nom, u.prenom
    ''', (date_fin_mois, date_debut_mois)).fetchall()

    return salaries, date_debut_mois, date_fin_mois


def _get_donnees_prepa(conn, salaries, mois, annee, date_debut_mois, date_fin_mois):
    """Construit les donnees de la grille prepa paie."""
    grille = []

    # Recuperer les statuts traite
    statuts_rows = conn.execute('''
        SELECT user_id, traite FROM prepa_paie_statut
        WHERE mois = %s AND annee = %s
    ''', (mois, annee)).fetchall()
    statuts = {r['user_id']: r['traite'] for r in statuts_rows}

    # Recuperer les variables paie du mois
    vp_rows = conn.execute('''
        SELECT * FROM variables_paie
        WHERE mois = %s AND annee = %s
    ''', (mois, annee)).fetchall()
    variables = {r['user_id']: dict(r) for r in vp_rows}

    for sal in salaries:
        uid = sal['id']

        # Contrats actifs sur le mois
        contrats = conn.execute('''
            SELECT id, type_contrat, date_debut, date_fin, forfait, nbr_jours,
                   fichier_path, fichier_nom
            FROM contrats
            WHERE user_id = %s
            AND date_debut <= %s
            AND (date_fin IS NULL OR date_fin >= %s)
            ORDER BY date_debut DESC
        ''', (uid, date_fin_mois, date_debut_mois)).fetchall()

        # Variables de paie
        vp = variables.get(uid, {})

        # Absences du mois (hors recuperations)
        placeholders = ','.join('%s' for _ in MOTIFS_ABSENCE_PAIE)
        absences = conn.execute(f'''
            SELECT id, motif, date_debut, date_fin, date_reprise, commentaire, jours_ouvres,
                   justificatif_path
            FROM absences
            WHERE user_id = %s
            AND motif IN ({placeholders})
            AND date_debut <= %s
            AND date_fin >= %s
            ORDER BY date_debut
        ''', (uid, *MOTIFS_ABSENCE_PAIE, date_fin_mois, date_debut_mois)).fetchall()

        grille.append({
            'user_id': uid,
            'nom': sal['nom'],
            'prenom': sal['prenom'],
            'secteur': sal['secteur_nom'],
            'traite': statuts.get(uid, 0),
            'contrats': [dict(c) for c in contrats],
            'mutuelle': vp.get('mutuelle', 0),
            'nb_enfants': vp.get('nb_enfants', 0),
            'heures_reelles': vp.get('heures_reelles'),
            'heures_supps': vp.get('heures_supps'),
            'transport': vp.get('transport', 0),
            'acompte': vp.get('acompte', 0),
            'saisie_salaire': vp.get('saisie_salaire', 0),
            'pret_avance': vp.get('pret_avance', 0),
            'autres_regularisation': vp.get('autres_regularisation', 0),
            'commentaire': vp.get('commentaire', ''),
            'absences': [dict(a) for a in absences],
        })

    return grille


@prepa_paie_bp.route('/prepa_paie', methods=['GET'])
@login_required
def prepa_paie():
    """Page principale : grille de preparation de paie mensuelle."""
    if not _peut_acceder_prepa_paie():
        flash("Acces non autorise.", 'error')
        return redirect(url_for('dashboard_bp.dashboard'))

    now = datetime.now()
    mois = request.args.get('mois', now.month, type=int)
    annee = request.args.get('annee', now.year, type=int)

    if mois < 1:
        mois = 12
        annee -= 1
    elif mois > 12:
        mois = 1
        annee += 1

    conn = get_db()

    salaries, date_debut_mois, date_fin_mois = _get_salaries_avec_contrat_actif(conn, mois, annee)
    grille = _get_donnees_prepa(conn, salaries, mois, annee, date_debut_mois, date_fin_mois)

    conn.close()

    # Navigation mois
    if mois == 1:
        prev_mois, prev_annee = 12, annee - 1
    else:
        prev_mois, prev_annee = mois - 1, annee

    if mois == 12:
        next_mois, next_annee = 1, annee + 1
    else:
        next_mois, next_annee = mois + 1, annee

    return render_template('prepa_paie.html',
                           grille=grille,
                           mois=mois,
                           annee=annee,
                           nom_mois=NOMS_MOIS[mois],
                           prev_mois=prev_mois,
                           prev_annee=prev_annee,
                           next_mois=next_mois,
                           next_annee=next_annee)


@prepa_paie_bp.route('/prepa_paie/traiter', methods=['POST'])
@login_required
def enregistrer_statut():
    """Enregistrer les cases 'traite' du mois."""
    if not _peut_acceder_prepa_paie():
        flash("Acces non autorise.", 'error')
        return redirect(url_for('dashboard_bp.dashboard'))

    mois = request.form.get('mois', type=int)
    annee = request.form.get('annee', type=int)
    user_ids = request.form.getlist('user_ids', type=int)

    if not mois or not annee:
        flash("Mois ou annee invalide.", 'error')
        return redirect(url_for('prepa_paie_bp.prepa_paie'))

    conn = get_db()
    try:
        for uid in user_ids:
            traite = 1 if request.form.get(f'traite_{uid}') else 0

            existing = conn.execute(
                'SELECT id FROM prepa_paie_statut WHERE user_id = %s AND mois = %s AND annee = %s',
                (uid, mois, annee)
            ).fetchone()

            if existing:
                conn.execute('''
                    UPDATE prepa_paie_statut
                    SET traite = %s, traite_par = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                ''', (traite, session['user_id'], existing['id']))
            else:
                conn.execute('''
                    INSERT INTO prepa_paie_statut (user_id, mois, annee, traite, traite_par)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (uid, mois, annee, traite, session['user_id']))

        conn.commit()
        flash(f"Statuts enregistres pour {NOMS_MOIS[mois]} {annee}.", 'success')
    except Exception as e:
        flash(f"Erreur : {str(e)}", 'error')
    finally:
        conn.close()

    return redirect(url_for('prepa_paie_bp.prepa_paie', mois=mois, annee=annee))


@prepa_paie_bp.route('/prepa_paie/export_excel')
@login_required
def export_excel():
    """Exporte la preparation de paie du mois en fichier Excel."""
    if not _peut_acceder_prepa_paie():
        flash("Acces non autorise.", 'error')
        return redirect(url_for('dashboard_bp.dashboard'))

    now = datetime.now()
    mois = request.args.get('mois', now.month, type=int)
    annee = request.args.get('annee', now.year, type=int)

    conn = get_db()
    salaries, date_debut_mois, date_fin_mois = _get_salaries_avec_contrat_actif(conn, mois, annee)
    grille = _get_donnees_prepa(conn, salaries, mois, annee, date_debut_mois, date_fin_mois)
    conn.close()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash("Le module openpyxl n'est pas installe. Export Excel indisponible.", 'error')
        return redirect(url_for('prepa_paie_bp.prepa_paie', mois=mois, annee=annee))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Prepa Paie {NOMS_MOIS[mois]} {annee}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    traite_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    # En-tetes
    headers = [
        'Traite', 'Nom', 'Prenom', 'Secteur',
        'Contrat (type)', 'Date debut', 'Date fin',
        'Forfait', 'Nbr. Jours',
        'Mutuelle', 'Enfants', 'H. reelles', 'H. supps',
        'Transport', 'Acompte',
        'Saisie/Salaire', 'Pret/Avance', 'Autres regul.', 'Commentaire',
        'Absences'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Donnees
    row_num = 2
    for item in grille:
        # Contrat principal (premier de la liste)
        contrat = item['contrats'][0] if item['contrats'] else {}

        # Absences sous forme texte
        absences_txt = ""
        for ab in item['absences']:
            absences_txt += f"{ab['motif']}: {ab['date_debut']} -> {ab['date_fin']}"
            if ab.get('date_reprise'):
                absences_txt += f" (reprise: {ab['date_reprise']})"
            if ab.get('commentaire'):
                absences_txt += f" [{ab['commentaire']}]"
            absences_txt += "\n"

        row_data = [
            'Oui' if item['traite'] else 'Non',
            item['nom'],
            item['prenom'],
            item['secteur'],
            contrat.get('type_contrat', ''),
            contrat.get('date_debut', ''),
            contrat.get('date_fin', ''),
            contrat.get('forfait', ''),
            contrat.get('nbr_jours', ''),
            'Oui' if item['mutuelle'] else 'Non',
            item['nb_enfants'],
            item['heures_reelles'] if item['heures_reelles'] else '',
            item['heures_supps'] if item['heures_supps'] else '',
            item['transport'] if item['transport'] else '',
            item['acompte'] if item['acompte'] else '',
            item['saisie_salaire'] if item['saisie_salaire'] else '',
            item['pret_avance'] if item['pret_avance'] else '',
            item['autres_regularisation'] if item['autres_regularisation'] else '',
            item['commentaire'] or '',
            absences_txt.strip(),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Colorer les lignes traitees
        if item['traite']:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = traite_fill

        # Si plusieurs contrats, ajouter des lignes supplementaires
        for extra_contrat in item['contrats'][1:]:
            row_num += 1
            ws.cell(row=row_num, column=1, value='').border = thin_border
            ws.cell(row=row_num, column=2, value='').border = thin_border
            ws.cell(row=row_num, column=3, value='').border = thin_border
            ws.cell(row=row_num, column=4, value='').border = thin_border
            ws.cell(row=row_num, column=5, value=extra_contrat.get('type_contrat', '')).border = thin_border
            ws.cell(row=row_num, column=6, value=extra_contrat.get('date_debut', '')).border = thin_border
            ws.cell(row=row_num, column=7, value=extra_contrat.get('date_fin', '')).border = thin_border
            ws.cell(row=row_num, column=8, value=extra_contrat.get('forfait', '')).border = thin_border
            ws.cell(row=row_num, column=9, value=extra_contrat.get('nbr_jours', '')).border = thin_border
            for c in range(10, len(headers) + 1):  # Colonnes après contrat
                ws.cell(row=row_num, column=c, value='').border = thin_border

        row_num += 1

    # Ajuster largeurs de colonnes
    col_widths = [8, 15, 15, 15, 12, 12, 12, 12, 10, 10, 8, 10, 10, 10, 10, 12, 12, 12, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Ecrire le fichier en memoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nom_fichier = f"prepa_paie_{NOMS_MOIS[mois]}_{annee}.xlsx"
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={nom_fichier}'
    return response
